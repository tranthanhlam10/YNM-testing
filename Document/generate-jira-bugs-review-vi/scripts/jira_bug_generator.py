#!/usr/bin/env python3
"""Xem trước hoặc tạo bug Jira Cloud từ Sheet, file hay bug nhập qua chat."""

from __future__ import annotations

import argparse
import base64
import csv
import io
import json
import os
import re
import sys
import unicodedata
import urllib.error
import urllib.request
from collections.abc import Iterable
from pathlib import Path
from typing import Any

ALIASES = {
    "test_case_id": ["test case id", "case id", "tc id", "test id", "id", "testcase id", "ma testcase", "ma test case"],
    "title": ["test case", "test case name", "test name", "name", "title", "scenario", "testcase", "ten testcase", "ten test case", "ten chuc nang", "kich ban"],
    "status": ["status", "result", "execution status", "test result", "execution result", "ket qua chay", "ket qua test", "trang thai"],
    "steps": ["steps", "test steps", "reproduction steps", "steps to reproduce", "procedure", "cac buoc thuc hien", "buoc thuc hien"],
    "expected": ["expected", "expected result", "expected outcome", "expected behavior", "ket qua mong doi", "mong doi"],
    "actual": ["actual", "actual result", "actual outcome", "observed result", "actual behavior", "ket qua thuc te", "thuc te"],
    "environment": ["environment", "env", "test environment", "platform", "moi truong", "moi truong test"],
    "preconditions": ["preconditions", "pre conditions", "pre condition", "prerequisites", "precondition", "dieu kien tien quyet", "dieu kien truoc"],
    "severity": ["severity", "priority", "impact", "muc do", "do uu tien"],
    "evidence": ["evidence", "attachment", "attachments", "screenshot", "log", "video", "bang chung", "tep dinh kem"],
    "component": ["component", "module", "module feature", "feature", "area", "chuc nang", "phan he"],
    "test_data": ["test data", "data", "du lieu test", "du lieu kiem thu"],
    "remarks": ["remarks", "remark", "notes", "note", "ghi chu"],
    "test_type": ["test type", "type", "loai test", "loai kiem thu"],
    "assigned_to": ["assigned to", "tester", "owner", "nguoi thuc hien", "nguoi phu trach"],
    "bug_id": ["bug id", "jira id", "jira key", "issue key", "link jira", "jira url", "ma bug"],
    "bug_status": ["bug status", "jira status", "issue status", "trang thai bug"],
    "bug_summary": ["bug summary", "bug title", "jira summary", "defect title", "summary", "tieu de bug"],
    "ready_to_jira": ["ready to jira", "ready to push", "push jira", "log jira", "create jira", "san sang tao jira"],
    "source_type": ["source type", "bug source", "nguon bug", "nguon nhap"],
    "selection_reason": ["selection reason", "candidate reason", "ly do chon", "ly do tao bug"],
    "root_cause_label": ["root cause", "root cause label", "rc label", "nguyen nhan goc"],
    "system_labels": ["system label", "system labels", "technical system", "he thong ky thuat"],
    "flow_labels": ["flow label", "flow labels", "technical flow", "luong ky thuat"],
    "jira_labels": ["jira labels", "labels", "label jira"],
}

PRIORITY_MAP = {
    "blocker": "Highest",
    "critical": "Highest",
    "highest": "Highest",
    "high": "High",
    "major": "High",
    "medium": "Medium",
    "normal": "Medium",
    "minor": "Low",
    "low": "Low",
    "lowest": "Lowest",
    "trivial": "Lowest",
    "nghiem trong": "Highest",
    "rat cao": "Highest",
    "cao": "High",
    "trung binh": "Medium",
    "thap": "Low",
}

CORE_JIRA_FIELDS = {"project", "summary", "issuetype", "description", "labels"}
CANONICAL_FIELDS = set(ALIASES)
MAX_CREATE_BATCH = 10
SELECTION_MODES = {"status", "ready", "all", "candidates"}
AMBIGUOUS_ACTUAL_MARKERS = (
    "can confirm",
    "chua check",
    "cho kiem tra",
    "cho confirm",
    "can kiem tra lai",
    "khong ro",
    "co ve",
    "dang doi dev",
    "doi dev fix",
    "cho dev fix",
    "need dev fix",
    "tbc",
)
CONTRADICTORY_ACTUAL_MARKERS = (
    "da dung yeu cau",
    "hoat dong binh thuong",
    "dung nhu mong doi",
)

ROOT_CAUSE_LABELS = {
    "rc-requirement",
    "rc-design-db",
    "rc-design-system",
    "rc-design-nonfunctional",
    "rc-logic",
    "rc-validation",
    "rc-data",
    "rc-config",
    "rc-integration",
    "rc-release",
    "rc-process",
    "rc-external-api",
    "rc-infra",
    "rc-assumption",
    "rc-document",
}
SYSTEM_LABELS = {
    "sys-crawling-auto",
    "sys-crawling-manual",
    "sys-crawling-adhoc",
    "sys-transform",
    "sys-ai",
    "sys-frontend",
    "sys-api",
    "sys-db",
    "sys-security",
    "sys-performance",
    "sys-infra",
}
TEST_TYPE_LABELS = {
    "test-functional",
    "test-regression",
    "test-boundary",
    "test-negative",
    "test-integration",
    "test-exploratory",
}
FLOW_LABELS = {
    "flow-source-load",
    "flow-source-build",
    "flow-auth",
    "flow-fetch",
    "flow-transform",
    "flow-pagination",
    "flow-pusher",
    "flow-updater",
    "flow-token-proxy-manager",
}
LIFECYCLE_LABELS = {"lc-reopen"}
TEST_TYPE_MAP = {
    "functional": "test-functional",
    "function": "test-functional",
    "chuc nang": "test-functional",
    "regression": "test-regression",
    "hoi quy": "test-regression",
    "boundary": "test-boundary",
    "boundary value": "test-boundary",
    "edge": "test-boundary",
    "edge case": "test-boundary",
    "gia tri bien": "test-boundary",
    "negative": "test-negative",
    "invalid data": "test-negative",
    "du lieu khong hop le": "test-negative",
    "integration": "test-integration",
    "tich hop": "test-integration",
    "exploratory": "test-exploratory",
    "explore": "test-exploratory",
    "tu do": "test-exploratory",
}


class InputError(ValueError):
    """Lỗi khi không thể diễn giải input một cách an toàn."""


class JiraError(RuntimeError):
    """Lỗi khi lời gọi HTTP tới Jira thất bại."""


def normalized(value: Any) -> str:
    text = str(value or "").replace("đ", "d").replace("Đ", "D")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value).strip()


def dedupe(values: Iterable[str]) -> list[str]:
    return list(dict.fromkeys(value for value in values if value))


def clean_label(value: Any) -> str:
    return re.sub(r"\s+", "-", clean(value).lower().replace("_", "-")).strip("-")


def split_labels(value: Any) -> list[str]:
    text = clean(value)
    if not text:
        return []
    if text.startswith("["):
        try:
            parsed = json.loads(text)
        except json.JSONDecodeError:
            parsed = None
        if isinstance(parsed, list):
            return dedupe(clean_label(item) for item in parsed)
    return dedupe(clean_label(item) for item in re.split(r"[,;\n]+", text))


def contains_marker(text: str, markers: Iterable[str]) -> bool:
    padded = f" {text} "
    return any(f" {normalized(marker)} " in padded for marker in markers)


def map_found_in_environment(value: str) -> str:
    text = normalized(value)
    if contains_marker(text, ("production", "prod", "san xuat")):
        return "Production"
    if contains_marker(text, ("staging", "stage")):
        return "Staging"
    if contains_marker(text, ("testing", "test env", "test environment", "qa env", "dev", "local")):
        return "Testing"
    return ""


def parse_json_rows(data: Any) -> list[dict[str, Any]]:
    if isinstance(data, dict):
        for key in ("testCases", "test_cases", "cases", "rows", "data"):
            if isinstance(data.get(key), list):
                data = data[key]
                break
    if not isinstance(data, list) or not all(isinstance(row, dict) for row in data):
        raise InputError("JSON phải là một mảng object hoặc chứa một array key được hỗ trợ")
    return data


def read_json(path: Path) -> list[dict[str, Any]]:
    try:
        data = json.loads(path.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError) as exc:
        raise InputError(f"Không thể đọc JSON input: {exc}") from exc
    return parse_json_rows(data)


def read_json_stdin() -> list[dict[str, Any]]:
    try:
        return parse_json_rows(json.load(sys.stdin))
    except json.JSONDecodeError as exc:
        raise InputError(f"Không thể đọc JSON từ stdin: {exc}") from exc


def read_csv(path: Path) -> list[dict[str, Any]]:
    try:
        raw = path.read_text(encoding="utf-8-sig")
    except (OSError, UnicodeDecodeError) as exc:
        raise InputError(f"Không thể đọc CSV dưới dạng UTF-8: {exc}") from exc
    try:
        dialect = csv.Sniffer().sniff(raw[:8192], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    rows = list(csv.DictReader(io.StringIO(raw), dialect=dialect))
    if not rows:
        raise InputError("CSV không có dòng dữ liệu")
    return rows


def read_xlsx(path: Path, sheet: str | None) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise InputError("Đọc XLSX cần openpyxl: python3 -m pip install openpyxl") from exc
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise InputError(f"Không thể đọc XLSX input: {exc}") from exc
    if sheet and sheet not in workbook.sheetnames:
        raise InputError(f"Không tìm thấy worksheet '{sheet}'. Hiện có: {', '.join(workbook.sheetnames)}")
    worksheet = workbook[sheet] if sheet else workbook.active
    values = list(worksheet.iter_rows(values_only=True))
    header_index = next((i for i, row in enumerate(values) if any(clean(v) for v in row)), None)
    if header_index is None:
        raise InputError("Worksheet XLSX đang trống")
    headers = [clean(value) for value in values[header_index]]
    if not any(headers):
        raise InputError("Dòng header XLSX đang trống")
    rows: list[dict[str, Any]] = []
    for values_row in values[header_index + 1 :]:
        if not any(clean(value) for value in values_row):
            continue
        rows.append({headers[i]: value for i, value in enumerate(values_row) if i < len(headers) and headers[i]})
    if not rows:
        raise InputError("Worksheet XLSX không có dữ liệu sau header")
    return rows


def read_rows(path: Path, sheet: str | None) -> list[dict[str, Any]]:
    if not path.exists() or not path.is_file():
        raise InputError(f"File input không tồn tại: {path}")
    suffix = path.suffix.lower()
    if suffix == ".json":
        return read_json(path)
    if suffix in {".csv", ".tsv"}:
        return read_csv(path)
    if suffix == ".xlsx":
        return read_xlsx(path, sheet)
    raise InputError("Định dạng input không được hỗ trợ. Hãy dùng .csv, .tsv, .json hoặc .xlsx")


def load_object(path_text: str | None, label: str) -> dict[str, Any]:
    if not path_text:
        return {}
    path = Path(path_text).expanduser().resolve()
    try:
        data = json.loads(path.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError) as exc:
        raise InputError(f"Không thể đọc {label}: {exc}") from exc
    if not isinstance(data, dict):
        raise InputError(f"{label} phải chứa một JSON object")
    return data


def build_header_map(headers: Iterable[str], explicit: dict[str, Any]) -> dict[str, str]:
    header_lookup = {normalized(header): header for header in headers if clean(header)}
    mapping: dict[str, str] = {}
    unknown = sorted(set(explicit) - CANONICAL_FIELDS)
    if unknown:
        raise InputError(f"Field chuẩn không được hỗ trợ trong field map: {', '.join(unknown)}")
    for canonical, source in explicit.items():
        source_normalized = normalized(source)
        if source_normalized not in header_lookup:
            raise InputError(f"Không tìm thấy header nguồn được map cho {canonical}: {source}")
        mapping[canonical] = header_lookup[source_normalized]
    for canonical, aliases in ALIASES.items():
        if canonical in mapping:
            continue
        for alias in aliases:
            if normalized(alias) in header_lookup:
                mapping[canonical] = header_lookup[normalized(alias)]
                break
    return mapping


def adf_text(text: str) -> dict[str, Any]:
    return {"type": "text", "text": text}


def adf_paragraph(text: str) -> dict[str, Any]:
    content = [adf_text(text)] if text else []
    return {"type": "paragraph", "content": content}


def add_section(content: list[dict[str, Any]], heading: str, value: str) -> None:
    if not value:
        return
    content.append({"type": "heading", "attrs": {"level": 3}, "content": [adf_text(heading)]})
    lines = value.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    for line in lines:
        content.append(adf_paragraph(line))


def add_bullet_section(content: list[dict[str, Any]], heading: str, values: list[str]) -> None:
    items = [value for value in values if value]
    if not items:
        return
    content.append({"type": "heading", "attrs": {"level": 3}, "content": [adf_text(heading)]})
    content.append({
        "type": "bulletList",
        "content": [
            {"type": "listItem", "content": [adf_paragraph(value)]}
            for value in items
        ],
    })


def build_description(
    record: dict[str, str],
    source_row: int,
    source_url: str,
    source_kind: str,
    selection_reason: str,
    label_classification: dict[str, Any],
) -> dict[str, Any]:
    content: list[dict[str, Any]] = []
    add_section(content, "Điều kiện tiên quyết", record["preconditions"])
    add_section(content, "Các bước tái hiện", record["steps"])
    add_section(content, "Kết quả thực tế", record["actual"])
    add_section(content, "Kết quả mong đợi", record["expected"])
    add_section(content, "Dữ liệu kiểm thử", record["test_data"])
    add_section(content, "Môi trường", record["environment"])
    label_lines = [
        f"Found In Environment: {label_classification['found_in_environment'] or 'Chưa xác định'}",
        "Root Cause: " + (
            ", ".join(label_classification["root_cause"])
            if label_classification["root_cause"]
            else "Chưa xác định — cập nhật trước khi đóng bug"
        ),
        "System: " + (", ".join(label_classification["system"]) or "Chưa xác định"),
        "Test Type: " + (", ".join(label_classification["test_type"]) or "Chưa xác định"),
    ]
    if label_classification["flow"]:
        label_lines.append("Flow: " + ", ".join(label_classification["flow"]))
    if label_classification["lifecycle"]:
        label_lines.append("Lifecycle: " + ", ".join(label_classification["lifecycle"]))
    add_bullet_section(content, "Phân loại label", label_lines)
    add_section(content, "Bằng chứng", record["evidence"] or "Chưa có bằng chứng được cung cấp.")
    add_section(content, "Ghi chú", record["remarks"])

    source_lines = []
    effective_source = record["source_type"] or source_kind
    source_lines.append(f"Nguồn nhập: {effective_source}")
    source_lines.append(f"Test case: {record['test_case_id'] or 'Không gắn với test case nào'}")
    if record["title"]:
        source_lines.append(f"Kịch bản kiểm thử: {record['title']}")
    if record["component"]:
        source_lines.append(f"Module/Feature: {record['component']}")
    if record["severity"]:
        source_lines.append(f"Priority nguồn: {record['severity']}")
    if record["test_type"]:
        source_lines.append(f"Loại kiểm thử: {record['test_type']}")
    if record["assigned_to"]:
        source_lines.append(f"Người thực hiện nguồn: {record['assigned_to']}")
    source_lines.append(f"Lý do chọn candidate: {record['selection_reason'] or selection_reason}")
    if effective_source != "chat":
        source_lines.append(f"Dòng nguồn: {source_row}")
    if source_url:
        source_lines.append(f"URL nguồn: {source_url}")
    if record["bug_status"]:
        source_lines.append(f"Trạng thái bug nguồn: {record['bug_status']}")
    if record["status"]:
        source_lines.append(f"Trạng thái test nguồn: {record['status']}")
    add_bullet_section(content, "Thông tin nguồn", source_lines)
    return {"type": "doc", "version": 1, "content": content}


def strip_generic_actual_prefix(value: str) -> str:
    text = re.sub(r"\s+", " ", value).strip()
    patterns = (
        r"^(?:hiện tại\s+)?(?:đang\s+)?(?:bị\s+)?(?:bug|lỗi)\s*[:\-]?\s*",
        r"^(?:actual(?: result)?|thực tế)\s*[:\-]\s*",
    )
    for pattern in patterns:
        text = re.sub(pattern, "", text, flags=re.IGNORECASE).strip()
    text = re.sub(r"\bredis\b", "Redis", text, flags=re.IGNORECASE)
    text = re.sub(r"\bapi\b", "API", text, flags=re.IGNORECASE)
    text = re.sub(r"\bui\b", "UI", text, flags=re.IGNORECASE)
    return text[:1].upper() + text[1:] if text else ""


def strip_test_title_tags(value: str) -> str:
    return re.sub(r"^(?:\s*\[[^\]]+\])+\s*", "", value).strip()


def derive_summary(record: dict[str, str]) -> str:
    supplied_summary = record["bug_summary"]
    supplied_normalized = normalized(supplied_summary)
    if supplied_normalized.startswith(("kiem tra", "verify", "validate", "test ")):
        supplied_summary = ""
    body = strip_generic_actual_prefix(supplied_summary or record["actual"])
    if not body:
        body = strip_test_title_tags(record["title"]) or "Hành vi lỗi cần làm rõ"
    component = record["component"].strip()
    component = re.sub(r"^\[([^\]]+)\]\s*", r"\1 - ", component).strip(" -")
    if component and not body.startswith("["):
        body = f"[{component}] {body}"
    return body


def safe_summary(summary_text: str) -> str:
    summary = re.sub(r"\s+", " ", summary_text).strip()
    if len(summary) <= 255:
        return summary
    return summary[:252].rstrip() + "..."


def canonical_record(row: dict[str, Any], mapping: dict[str, str]) -> dict[str, str]:
    return {field: clean(row.get(mapping[field], "")) if field in mapping else "" for field in ALIASES}


def infer_system_labels(record: dict[str, str]) -> list[str]:
    primary = normalized(" ".join((
        record["bug_summary"],
        record["title"],
        record["steps"],
        record["expected"],
        record["actual"],
        record["test_data"],
        record["remarks"],
    )))
    text = primary or normalized(record["component"])
    labels: list[str] = []

    crawling = contains_marker(text, (
        "crawl", "crawling", "crawler", "scrape", "scraping", "thu thap du lieu nguon ngoai",
    ))
    if crawling:
        if contains_marker(text, ("adhoc", "one time", "tam thoi", "debug script", "script debug")):
            labels.append("sys-crawling-adhoc")
        elif contains_marker(text, ("manual crawl", "crawl manual", "trigger thu cong", "admin tool", "tool admin")):
            labels.append("sys-crawling-manual")
        else:
            labels.append("sys-crawling-auto")

    rules = (
        ("sys-transform", ("transform", "mapping", "map field", "normalize data", "chuan hoa du lieu")),
        ("sys-ai", ("ai", "ml", "model", "prompt", "inference", "scoring")),
        ("sys-frontend", (
            "ui", "ux", "giao dien", "hien thi", "client", "browser", "mobile", "responsive",
            "button", "filter", "panel", "sidebar", "modal", "dropdown", "chip", "badge",
        )),
        ("sys-api", ("api", "request", "response", "endpoint", "contract", "http", "status code", "timeout")),
        ("sys-db", ("database", "db", "schema", "index", "constraint", "migration", "query", "sql")),
        ("sys-security", (
            "security", "bao mat", "auth", "authentication", "authorization", "permission", "role",
            "phan quyen", "data exposure", "lo du lieu",
        )),
        ("sys-performance", (
            "performance", "hieu nang", "slow", "latency", "high load", "memory", "cham",
        )),
        ("sys-infra", ("infra", "server", "network", "ci cd", "cloud resource", "ha tang")),
    )
    for label, markers in rules:
        if contains_marker(text, markers):
            labels.append(label)
    return dedupe(labels)


def infer_flow_labels(record: dict[str, str]) -> list[str]:
    text = normalized(" ".join((
        record["bug_summary"], record["title"], record["steps"], record["actual"], record["remarks"],
    )))
    rules = (
        ("flow-source-load", ("load source config", "source config fail", "sai config url", "sai config param")),
        ("flow-source-build", ("build url", "build request", "mapping param dau vao", "map param dau vao")),
        ("flow-auth", ("expired credential", "token het han", "sai token", "thieu token", "sai proxy", "thieu proxy", "sai cookie")),
        ("flow-fetch", ("call api fail", "api timeout", "response sai format", "raw response")),
        ("flow-transform", ("mapping sai field", "map sai field", "transform sai", "xu ly data sai")),
        ("flow-pagination", ("next page", "has more", "pagination", "phan trang", "miss data", "duplicate data")),
        ("flow-pusher", ("insert fail", "insert loi", "duplicate trong db", "missing data trong db")),
        ("flow-updater", ("update fail sau crawling", "update sai sau crawling", "update thieu sau crawling")),
        ("flow-token-proxy-manager", ("token proxy manager", "service token", "service proxy")),
    )
    return [label for label, markers in rules if contains_marker(text, markers)]


def classify_team_labels(
    record: dict[str, str],
    provided_labels: list[str],
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    warnings: list[dict[str, Any]] = []

    def add_warning(code: str, message: str, blocking: bool) -> None:
        warnings.append({"code": code, "message": message, "blocking": blocking})

    raw_labels = dedupe([*provided_labels, *split_labels(record["jira_labels"])])
    root_values = dedupe([
        *split_labels(record["root_cause_label"]),
        *(label for label in raw_labels if label.startswith("rc-")),
    ])
    system_values = dedupe([
        *split_labels(record["system_labels"]),
        *(label for label in raw_labels if label.startswith("sys-")),
    ])
    flow_values = dedupe([
        *split_labels(record["flow_labels"]),
        *(label for label in raw_labels if label.startswith("flow-")),
    ])
    test_values = dedupe(label for label in raw_labels if label.startswith("test-"))
    lifecycle_values = dedupe(label for label in raw_labels if label.startswith("lc-"))

    invalid_root = [label for label in root_values if label not in ROOT_CAUSE_LABELS]
    roots = [label for label in root_values if label in ROOT_CAUSE_LABELS]
    if invalid_root:
        add_warning("invalid_root_cause_label", f"Root Cause label không thuộc taxonomy: {', '.join(invalid_root)}", True)
    if len(roots) > 1:
        add_warning("multiple_root_cause_labels", "Mỗi bug chỉ được có đúng một Root Cause label", True)
        roots = []
    if not roots and not invalid_root and len(root_values) <= 1:
        add_warning(
            "root_cause_pending",
            "Chưa xác định Root Cause; được phép để trống lúc tạo nhưng phải cập nhật trước khi đóng bug",
            False,
        )

    invalid_system = [label for label in system_values if label not in SYSTEM_LABELS]
    systems = [label for label in system_values if label in SYSTEM_LABELS]
    if invalid_system:
        add_warning("invalid_system_label", f"System label không thuộc taxonomy: {', '.join(invalid_system)}", True)
    if not system_values:
        systems = infer_system_labels(record)
    if not systems:
        add_warning("missing_system_label", "Không xác định được System label từ hành vi bug", True)

    invalid_flow = [label for label in flow_values if label not in FLOW_LABELS]
    flows = [label for label in flow_values if label in FLOW_LABELS]
    if invalid_flow:
        add_warning("invalid_flow_label", f"Flow label không thuộc taxonomy: {', '.join(invalid_flow)}", True)
    if not flow_values:
        flows = infer_flow_labels(record)

    invalid_test = [label for label in test_values if label not in TEST_TYPE_LABELS]
    tests = [label for label in test_values if label in TEST_TYPE_LABELS]
    if invalid_test:
        add_warning("invalid_test_type_label", f"Test Type label không thuộc taxonomy: {', '.join(invalid_test)}", True)
    source_test_type = normalized(record["test_type"])
    if source_test_type:
        if source_test_type in TEST_TYPE_MAP:
            tests.append(TEST_TYPE_MAP[source_test_type])
        elif clean_label(record["test_type"]) in TEST_TYPE_LABELS:
            tests.append(clean_label(record["test_type"]))
        elif not test_values:
            add_warning("unmapped_test_type", f"Không map được TEST TYPE nguồn: {record['test_type']}", True)
    tests = dedupe(tests)
    if len(tests) > 1:
        add_warning("multiple_test_type_labels", "Mỗi bug chỉ được có đúng một Test Type label", True)
        tests = []
    elif not tests:
        add_warning("missing_test_type_label", "Chưa xác định được Test Type theo hoạt động test thực tế", False)

    invalid_lifecycle = [label for label in lifecycle_values if label not in LIFECYCLE_LABELS]
    lifecycle = [label for label in lifecycle_values if label in LIFECYCLE_LABELS]
    if invalid_lifecycle:
        add_warning("invalid_lifecycle_label", f"Lifecycle label không thuộc taxonomy: {', '.join(invalid_lifecycle)}", True)
    if normalized(record["bug_status"]) in {"reopen", "reopened", "mo lai"}:
        lifecycle.append("lc-reopen")
    lifecycle = dedupe(lifecycle)

    taxonomy = ROOT_CAUSE_LABELS | SYSTEM_LABELS | TEST_TYPE_LABELS | FLOW_LABELS | LIFECYCLE_LABELS
    operational = [label for label in raw_labels if label not in taxonomy]
    found_in_environment = map_found_in_environment(record["environment"])
    if record["environment"] and not found_in_environment:
        add_warning(
            "unmapped_found_in_environment",
            "Môi trường không map được sang Found In Environment: Testing, Staging hoặc Production",
            True,
        )

    return {
        "found_in_environment": found_in_environment,
        "root_cause": roots,
        "system": dedupe(systems),
        "test_type": tests,
        "flow": dedupe(flows),
        "lifecycle": lifecycle,
        "operational": operational,
    }, warnings


def evaluate_quality(record: dict[str, str]) -> tuple[str, list[dict[str, Any]]]:
    warnings: list[dict[str, Any]] = []

    def add_warning(code: str, message: str, blocking: bool) -> None:
        warnings.append({"code": code, "message": message, "blocking": blocking})

    actual = normalized(record["actual"])
    expected = normalized(record["expected"])
    bug_summary = normalized(record["bug_summary"])
    if any(marker in actual for marker in AMBIGUOUS_ACTUAL_MARKERS):
        add_warning("actual_needs_confirmation", "ACTUAL RESULT chứa nội dung chưa được xác nhận rõ", True)
    vague_actuals = {
        "loi",
        "bi loi",
        "dang loi",
        "hien tai dang loi",
        "hien tai bi loi",
        "khong dung",
        "sai",
    }
    if actual in vague_actuals or len(actual) < 10:
        add_warning("actual_too_vague", "ACTUAL RESULT quá ngắn hoặc chưa mô tả hành vi quan sát được", True)
    if actual and expected and actual == expected:
        add_warning("expected_equals_actual", "EXPECTED RESULT và ACTUAL RESULT giống nhau", True)
    if any(marker in actual for marker in CONTRADICTORY_ACTUAL_MARKERS):
        add_warning("candidate_actual_conflict", "Candidate được chọn nhưng ACTUAL RESULT cho biết hệ thống đang đúng", True)
    if bug_summary.startswith(("kiem tra", "verify", "validate", "test ")):
        add_warning("bug_summary_is_test_intent", "BUG SUMMARY đang mô tả mục tiêu kiểm thử; skill sẽ dùng ACTUAL RESULT thay thế", False)
    if not record["steps"]:
        add_warning("missing_steps", "Thiếu TEST STEPS hoặc bước tái hiện tương đương", True)
    if not record["environment"]:
        add_warning("missing_environment", "Thiếu môi trường kiểm thử", True)
    if not record["severity"]:
        add_warning("missing_priority", "Thiếu priority/severity; skill sẽ không tự suy đoán", True)
    elif normalized(record["severity"]) not in PRIORITY_MAP:
        add_warning("unsupported_priority", "Priority/severity nguồn chưa map được sang Jira", True)
    if not record["evidence"]:
        add_warning("missing_evidence", "Không có đường dẫn hoặc URL bằng chứng", False)

    review_state = "needs_clarification" if any(item["blocking"] for item in warnings) else "ready"
    return review_state, warnings


def build_preview(
    rows: list[dict[str, Any]],
    project: str,
    issue_type: str,
    include_statuses: set[str],
    labels: list[str],
    field_map: dict[str, Any],
    extra_fields: dict[str, Any],
    source_url: str = "",
    selection_mode: str = "ready",
    ready_values: set[str] | None = None,
    source_kind: str = "file",
    found_in_environment_field: str = "",
) -> dict[str, Any]:
    if not rows:
        raise InputError("Input không có dòng dữ liệu")
    mapping = build_header_map((key for row in rows for key in row), field_map)
    if selection_mode not in SELECTION_MODES:
        raise InputError(f"Selection mode không hợp lệ: {selection_mode}")
    if selection_mode == "status" and "status" not in mapping:
        raise InputError("Selection mode status cần cột status/result; hãy đổi mode hoặc cung cấp --field-map")
    if selection_mode == "ready" and "ready_to_jira" not in mapping:
        raise InputError("Selection mode ready cần cột READY TO JIRA/Ready to Push")
    forbidden = sorted(CORE_JIRA_FIELDS.intersection(extra_fields))
    if forbidden:
        raise InputError(f"Extra fields không được ghi đè field Jira cốt lõi: {', '.join(forbidden)}")

    drafts: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    all_quality_warnings: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    seen_fingerprints: set[str] = set()
    skipped_selection = 0
    skipped_existing_bug_id = 0
    effective_ready_values = ready_values or {"yes", "ready", "true", "1"}

    for source_row, row in enumerate(rows, start=2):
        record = canonical_record(row, mapping)
        status = normalized(record["status"])
        ready_value = normalized(record["ready_to_jira"])
        if selection_mode == "status":
            selected = status in include_statuses
            computed_selection_reason = f"status:{record['status']}"
        elif selection_mode == "ready":
            selected = ready_value in effective_ready_values
            computed_selection_reason = f"ready:{record['ready_to_jira']}"
        else:
            selected = True
            computed_selection_reason = (
                "chưa có tín hiệu chọn - chỉ preview"
                if selection_mode == "candidates"
                else "chat hoặc dòng đã được chọn trước"
            )
        if not selected:
            skipped_selection += 1
            continue
        if record["bug_id"]:
            skipped_existing_bug_id += 1
            skipped.append({
                "source_row": source_row,
                "test_case_id": record["test_case_id"],
                "bug_id": record["bug_id"],
                "review_state": "skip_existing",
                "reason": "đã có BUG ID",
            })
            continue
        identity = normalized(record["test_case_id"])
        if identity and identity in seen_ids:
            errors.append({"source_row": source_row, "test_case_id": record["test_case_id"], "review_state": "invalid", "error": "trùng test-case ID"})
            continue
        missing = []
        if not record["test_case_id"] and not record["title"] and not record["bug_summary"]:
            missing.append("summary hoặc title")
        if not record["expected"]:
            missing.append("expected")
        if not record["actual"]:
            missing.append("actual")
        if missing:
            errors.append({"source_row": source_row, "test_case_id": record["test_case_id"], "review_state": "invalid", "error": f"thiếu dữ liệu bắt buộc: {', '.join(missing)}"})
            continue
        if identity:
            seen_ids.add(identity)

        fingerprint = normalized(" | ".join((record["component"], derive_summary(record), record["actual"])))
        if fingerprint and fingerprint in seen_fingerprints:
            errors.append({
                "source_row": source_row,
                "test_case_id": record["test_case_id"],
                "review_state": "invalid",
                "error": "trùng nội dung bug trong cùng input",
            })
            continue
        if fingerprint:
            seen_fingerprints.add(fingerprint)

        review_state, quality_warnings = evaluate_quality(record)
        label_classification, label_warnings = classify_team_labels(record, labels)
        quality_warnings.extend(label_warnings)
        if any(item["blocking"] for item in quality_warnings):
            review_state = "needs_clarification"
        for warning in quality_warnings:
            all_quality_warnings.append({
                "source_row": source_row,
                "test_case_id": record["test_case_id"],
                **warning,
            })

        fields: dict[str, Any] = {
            "project": {"key": project},
            "summary": safe_summary(derive_summary(record)),
            "issuetype": {"name": issue_type},
            "description": build_description(
                record,
                source_row,
                source_url,
                source_kind,
                computed_selection_reason,
                label_classification,
            ),
        }
        priority = PRIORITY_MAP.get(normalized(record["severity"]))
        if priority:
            fields["priority"] = {"name": priority}
        if found_in_environment_field and label_classification["found_in_environment"]:
            fields[found_in_environment_field] = {"value": label_classification["found_in_environment"]}
        issue_labels = [
            *label_classification["operational"],
            *label_classification["root_cause"],
            *label_classification["system"],
            *label_classification["test_type"],
            *label_classification["flow"],
            *label_classification["lifecycle"],
        ]
        issue_labels.append("linked-testcase" if record["test_case_id"] else "no-testcase")
        issue_labels = list(dict.fromkeys(issue_labels))
        if issue_labels:
            fields["labels"] = issue_labels
        fields.update(extra_fields)
        drafts.append({
            "source_row": source_row,
            "test_case_id": record["test_case_id"],
            "source_status": record["status"],
            "source_severity": record["severity"],
            "source_bug_status": record["bug_status"],
            "source_type": record["source_type"] or source_kind,
            "selection_reason": record["selection_reason"] or computed_selection_reason,
            "review_state": review_state,
            "found_in_environment": label_classification["found_in_environment"],
            "label_classification": label_classification,
            "quality_warnings": quality_warnings,
            "payload": {"fields": fields},
        })

    return {
        "project": project,
        "issue_type": issue_type,
        "selection_mode": selection_mode,
        "source_kind": source_kind,
        "header_mapping": mapping,
        "drafts": drafts,
        "errors": errors,
        "skipped": skipped,
        "quality_warnings": all_quality_warnings,
        "source_url": source_url or None,
        "found_in_environment_field": found_in_environment_field or None,
        "stats": {
            "input_rows": len(rows),
            "drafts": len(drafts),
            "validation_errors": len(errors),
            "skipped_by_selection": skipped_selection,
            "skipped_by_status": skipped_selection if selection_mode == "status" else 0,
            "skipped_existing_bug_id": skipped_existing_bug_id,
            "quality_warnings": len(all_quality_warnings),
            "ready": sum(1 for item in drafts if item["review_state"] == "ready"),
            "needs_clarification": sum(1 for item in drafts if item["review_state"] == "needs_clarification"),
            "invalid": len(errors),
            "skip_existing": skipped_existing_bug_id,
        },
    }


def jira_credentials() -> tuple[str, str, str]:
    base_url = os.environ.get("JIRA_BASE_URL", "").strip().rstrip("/")
    email = os.environ.get("JIRA_EMAIL", "").strip()
    token = os.environ.get("JIRA_API_TOKEN", "").strip()
    missing = [name for name, value in (("JIRA_BASE_URL", base_url), ("JIRA_EMAIL", email), ("JIRA_API_TOKEN", token)) if not value]
    if missing:
        raise InputError(f"Thiếu biến môi trường Jira: {', '.join(missing)}")
    if not base_url.startswith("https://"):
        raise InputError("JIRA_BASE_URL phải bắt đầu bằng https://")
    return base_url, email, token


def jira_request(method: str, path: str, payload: dict[str, Any] | None = None) -> tuple[int, dict[str, Any]]:
    base_url, email, token = jira_credentials()
    credentials = base64.b64encode(f"{email}:{token}".encode()).decode("ascii")
    body = json.dumps(payload).encode("utf-8") if payload is not None else None
    request = urllib.request.Request(
        f"{base_url}{path}",
        data=body,
        method=method,
        headers={
            "Accept": "application/json",
            "Content-Type": "application/json",
            "Authorization": f"Basic {credentials}",
            "User-Agent": "generate-jira-bugs-skill/1.0",
        },
    )
    try:
        with urllib.request.urlopen(request, timeout=30) as response:
            raw = response.read().decode("utf-8")
            return response.status, json.loads(raw) if raw else {}
    except urllib.error.HTTPError as exc:
        raw = exc.read().decode("utf-8", errors="replace")
        try:
            details = json.loads(raw)
        except json.JSONDecodeError:
            details = {"message": raw[:2000]}
        raise JiraError(f"Jira trả về HTTP {exc.code}: {json.dumps(details, ensure_ascii=False)}") from exc
    except urllib.error.URLError as exc:
        raise JiraError(f"Không thể kết nối Jira: {exc.reason}") from exc


def check_auth() -> dict[str, Any]:
    _, profile = jira_request("GET", "/rest/api/3/myself")
    return {
        "ok": True,
        "account_id": profile.get("accountId"),
        "display_name": profile.get("displayName"),
        "email": profile.get("emailAddress"),
    }


def create_issues(drafts: list[dict[str, Any]]) -> list[dict[str, Any]]:
    base_url, _, _ = jira_credentials()
    results = []
    for draft in drafts:
        summary = draft["payload"]["fields"]["summary"]
        try:
            status, data = jira_request("POST", "/rest/api/3/issue", draft["payload"])
            key = data.get("key")
            results.append({
                "ok": status == 201 and bool(key),
                "test_case_id": draft["test_case_id"],
                "summary": summary,
                "id": data.get("id"),
                "key": key,
                "url": f"{base_url}/browse/{key}" if key else None,
            })
        except (JiraError, InputError) as exc:
            results.append({
                "ok": False,
                "test_case_id": draft["test_case_id"],
                "summary": summary,
                "error": str(exc),
            })
    return results


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", help="Đường dẫn CSV, TSV, JSON hoặc XLSX; dùng - để đọc JSON stdin")
    parser.add_argument("--project", default=os.environ.get("JIRA_PROJECT_KEY"), help="Jira project key")
    parser.add_argument("--issue-type", default="Bug", help="Tên Jira issue type (mặc định: Bug)")
    parser.add_argument("--sheet", help="Tên worksheet XLSX; mặc định là worksheet active")
    parser.add_argument("--include-status", default="bug,failed,fail,error,errored,thất bại,lỗi", help="Các result được chọn, phân cách bằng dấu phẩy")
    parser.add_argument("--selection-mode", choices=sorted(SELECTION_MODES), default="ready", help="Cách chọn candidate: ready (mặc định), status, all hoặc candidates (preview-only)")
    parser.add_argument("--ready-values", default="yes,ready,true,1", help="Các giá trị READY TO JIRA được chọn")
    parser.add_argument("--source-kind", choices=("auto", "sheet", "file", "chat"), default="auto", help="Nguồn input để truy vết")
    parser.add_argument("--labels", default="generated-by-qc", help="Các Jira label bổ sung, phân cách bằng dấu phẩy")
    parser.add_argument("--field-map", help="File JSON map field chuẩn sang header nguồn")
    parser.add_argument("--extra-fields", help="JSON object chứa field Jira bổ sung")
    parser.add_argument(
        "--found-in-environment-field",
        default=os.environ.get("JIRA_FOUND_IN_ENVIRONMENT_FIELD", ""),
        help="Jira custom field ID cho Found In Environment, ví dụ customfield_12345",
    )
    parser.add_argument("--source-url", default="", help="URL Google Sheet hoặc test run nguồn")
    parser.add_argument("--output", help="Ghi kết quả JSON vào đường dẫn này; mặc định in ra stdout")
    parser.add_argument("--create", action="store_true", help="Tạo issue sau khi đã preview và được duyệt")
    parser.add_argument("--yes", action="store_true", help="Xác nhận bắt buộc khi dùng --create")
    parser.add_argument("--allow-quality-warnings", action="store_true", help="Cho phép tạo draft có cảnh báo chặn sau khi người dùng duyệt rõ")
    parser.add_argument("--check-auth", action="store_true", help="Kiểm tra credentials Jira mà không tạo issue")
    return parser.parse_args()


def write_result(result: dict[str, Any], output: str | None) -> None:
    rendered = json.dumps(result, ensure_ascii=False, indent=2)
    if output:
        path = Path(output).expanduser().resolve()
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(rendered + "\n", encoding="utf-8")
        print(f"Đã ghi kết quả vào {path}", file=sys.stderr)
    else:
        print(rendered)


def main() -> int:
    args = parse_args()
    try:
        if args.check_auth:
            write_result(check_auth(), args.output)
            return 0
        if not args.input:
            raise InputError("Bắt buộc có --input, trừ khi dùng --check-auth")
        if not args.project or not re.fullmatch(r"[A-Za-z][A-Za-z0-9_]*", args.project.strip()):
            raise InputError("Hãy cung cấp Jira project key hợp lệ bằng --project hoặc JIRA_PROJECT_KEY")
        if args.create and not args.yes:
            raise InputError("Tạo issue thật cần đồng thời --create và --yes sau khi người dùng xác nhận rõ")
        if args.create and args.project.strip().upper() in {"PREVIEW", "TBD"}:
            raise InputError("Project PREVIEW/TBD chỉ dùng để xem trước; hãy cung cấp Jira project key thật")
        if args.found_in_environment_field and not re.fullmatch(r"customfield_[0-9]+", args.found_in_environment_field):
            raise InputError("--found-in-environment-field phải có dạng customfield_<số>")
        if args.create and not args.found_in_environment_field:
            raise InputError(
                "Tạo issue thật cần --found-in-environment-field hoặc JIRA_FOUND_IN_ENVIRONMENT_FIELD "
                "để ghi custom field Found In Environment"
            )

        if args.input == "-":
            rows = read_json_stdin()
            input_label = "stdin"
        else:
            path = Path(args.input).expanduser().resolve()
            rows = read_rows(path, args.sheet)
            input_label = str(path)
        field_map = load_object(args.field_map, "field map")
        extra_fields = load_object(args.extra_fields, "extra fields")
        statuses = {normalized(value) for value in args.include_status.split(",") if normalized(value)}
        if not statuses:
            raise InputError("--include-status phải có ít nhất một giá trị")
        labels = [value.strip() for value in args.labels.split(",") if value.strip()]
        ready_values = {normalized(value) for value in args.ready_values.split(",") if normalized(value)}
        if not ready_values:
            raise InputError("--ready-values phải có ít nhất một giá trị")
        if args.source_kind == "auto":
            source_kind = "sheet" if args.source_url else ("chat" if args.input == "-" else "file")
        else:
            source_kind = args.source_kind
        preview = build_preview(
            rows,
            args.project.upper(),
            args.issue_type,
            statuses,
            labels,
            field_map,
            extra_fields,
            args.source_url,
            args.selection_mode,
            ready_values,
            source_kind,
            args.found_in_environment_field,
        )
        preview["input_file"] = input_label
        if args.create:
            if args.selection_mode == "candidates":
                raise InputError("Selection mode candidates chỉ dùng để preview; hãy chọn chính xác dòng rồi dùng mode all/ready/status")
            ready_drafts = [item for item in preview["drafts"] if item["review_state"] == "ready"]
            blocked = [item for item in preview["drafts"] if item["review_state"] == "needs_clarification"]
            if args.allow_quality_warnings:
                drafts_to_create = ready_drafts + blocked
            else:
                drafts_to_create = ready_drafts
            if not drafts_to_create:
                if blocked:
                    ids = ", ".join(item["test_case_id"] or f"dòng {item['source_row']}" for item in blocked)
                    raise InputError(
                        f"Không có draft READY để tạo; {len(blocked)} draft NEEDS_CLARIFICATION ({ids}). "
                        "Sửa dữ liệu hoặc xác nhận rõ trước khi dùng --allow-quality-warnings"
                    )
                raise InputError("Không có draft READY để tạo")
            if len(drafts_to_create) > MAX_CREATE_BATCH:
                raise InputError(f"Một batch chỉ được tạo tối đa {MAX_CREATE_BATCH} issue; hãy chia nhỏ và xin duyệt từng batch")
            preview["creation_skipped"] = [
                {
                    "source_row": item["source_row"],
                    "test_case_id": item["test_case_id"],
                    "summary": item["payload"]["fields"]["summary"],
                    "review_state": item["review_state"],
                    "reason": "needs_clarification",
                }
                for item in blocked
                if not args.allow_quality_warnings
            ]
            preview["creation_results"] = create_issues(drafts_to_create)
            preview["stats"]["created"] = sum(1 for item in preview["creation_results"] if item["ok"])
            preview["stats"]["create_failed"] = sum(1 for item in preview["creation_results"] if not item["ok"])
            preview["stats"]["skipped_needs_clarification"] = len(preview["creation_skipped"])
        write_result(preview, args.output)
        return 0 if not args.create or preview["stats"]["create_failed"] == 0 else 2
    except (InputError, JiraError) as exc:
        print(f"Lỗi: {exc}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
